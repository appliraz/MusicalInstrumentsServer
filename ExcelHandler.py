from openpyxl import Workbook, worksheet
from openpyxl.styles import Font, Alignment
from configurations import variablesService as vs
from fileService import getFileStream





def getDataAsExcelFileStream(complete_data: tuple):
    print("\n\n COMPLETE DATA \n\n")
    print(complete_data)
    sheet_no = 1
    wb = Workbook()
    if not complete_data:
        raise Exception("\nRequested to write empty data to file, cannot do that")
    for scraped_data, data_name in complete_data:
        try:
            wb = writeDataToWorkbook(scraped_data, data_name, sheet_no, wb)
            sheet_no += 1
        except Exception as e:
            print(e)
            continue
    wb = deleteUnnessacerySheets(wb)
    filestream = getFileStream(wb)
    return filestream


def writeDataToWorkbook(data: list, data_name: str, sheet_no: int, wb: Workbook):
    sheet_name = getSheetName(data_name, sheet_no)
    print(f'writing data to sheet {sheet_name}')
    ws = wb.create_sheet(sheet_name)
    ws = addHeadersToSheet(ws)
    ws = addDataToSheet(ws, data)
    ws = reformatSheet(ws)
    return wb


""" Utility Functions """

def reformatSheet(sheet):
    sheet = makeFirstRowBold(sheet)
    sheet = formatPriceColumn(sheet)
    sheet = adjustColumns(sheet)
    return sheet
   


def makeFirstRowBold(sheet):
    # make the first row bold
    print(f'reformating sheet')
    row = sheet[1]
    for cell in row:
        cell.font = Font(bold=True)
    return sheet


def formatPriceColumn(sheet):
    # make the price column in number format
    max_row = sheet.max_row
    for i in range(1, max_row+1):
        cell_index = "C" + str(i)
        try:
            sheet[cell_index].value = float(sheet[cell_index].value)
        except Exception as e:
            print(e)
            continue
    return sheet

def adjustColumns(sheet):
    max_row = sheet.max_row
    max_col = sheet.max_column
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 40
    for row in sheet.iter_rows(min_row=1, max_col=max_col, max_row=max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return sheet

def deleteUnnessacerySheets(wb: Workbook):
    sheets_to_delete = ['Sheet1', 'Sheet']
    for sheet in sheets_to_delete:
        if sheet in wb.sheetnames:
            del wb[sheet]
    return wb

def getProductRelevantParameters(product):
    link, model, price = product
    upperprice, lowerprice = price
    return (link, model, lowerprice)

def getSheetName(data_name: str, sheet_no: int):
    sheet_name = data_name + f" {str(sheet_no)}"
    return sheet_name

def addHeadersToSheet(sheet):
    headers = [vs.link_header, vs.name_header, vs.price_header]
    sheet.append(headers)
    return sheet

def addDataToSheet(sheet, data: list):
    for product in data:
        row = getProductRelevantParameters(product)
        sheet.append(row)
    return sheet